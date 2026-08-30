#!/usr/bin/env python3
"""Build the deduplicated S&P 500 valuation universe.

The paid Sharadar S&P 500 snapshot provides the source ticker and issuer
metadata.  An optional official SPY holdings workbook supplies the public
ticker, CUSIP, and portfolio weight used to choose a primary share class.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import pyarrow.dataset as ds


DEFAULT_JANSEN_ROOT = Path(
    os.environ.get(
        "JANSEN_SHARADAR_ROOT",
        Path.home() / "Documents/jansen_us_firm_replication/data/sharadar/parquet",
    )
)
DEFAULT_OUTPUT = Path("server/config/sp500-valuation-universe.json")
OFFICIAL_HOLDINGS_URL = (
    "https://www.ssga.com/library-content/products/fund-data/etfs/us/"
    "holdings-daily-us-en-spy.xlsx"
)


def valuation_profile(ticker: str, sector: str | None, industry: str | None) -> str:
    """Return a conservative, explicit valuation family for every issuer.

    Ticker-specific exceptions remain in the model code. This mapping is the
    required fallback for the rest of the index and deliberately has no
    mega-cap catch-all.
    """

    if ticker in {"CPAY", "FI", "FIS", "FISV", "GPN", "PYPL", "XYZ"}:
        return "payments_processor"

    sector = str(sector or "")
    industry = str(industry or "")
    if sector == "Financial Services":
        if industry.startswith("Banks -"):
            return "bank"
        if industry.startswith("Insurance -"):
            return "insurance"
        if industry == "Insurance Brokers":
            return "insurance_broker"
        if industry == "Credit Services":
            return "credit_services"
        if industry == "Financial Data & Stock Exchanges":
            return "information_services"
        if industry == "Capital Markets":
            return "capital_markets"
        return "asset_manager"
    if sector == "Real Estate":
        return "reit"
    if sector == "Utilities":
        return "power_utility"
    if sector == "Energy":
        return "energy_e_and_p" if industry == "Oil & Gas E&P" else "energy_infrastructure"
    if sector == "Basic Materials":
        return "materials"
    if sector == "Consumer Defensive":
        if industry == "Farm Products":
            return "commodity_merchant"
        return "consumer_staples"
    if sector == "Consumer Cyclical":
        if industry == "Internet Retail":
            return "platform_marketplace_reinvestment"
        return "consumer_cyclical"
    if sector == "Healthcare":
        if industry in {"Drug Manufacturers - General", "Drug Manufacturers - Specialty & Generic"}:
            return "biopharma"
        if industry == "Biotechnology":
            return "biopharma_growth"
        if industry in {"Medical Devices", "Medical Instruments & Supplies"}:
            return "mature_medtech"
        if industry == "Healthcare Plans":
            return "managed_care"
        if industry == "Medical Distribution":
            return "healthcare_distribution"
        return "healthcare_services"
    if sector == "Communication Services":
        if industry == "Internet Content & Information":
            return "platform_marketplace_reinvestment"
        if industry == "Electronic Gaming & Multimedia":
            return "interactive_entertainment"
        return "media_telecom"
    if sector == "Technology":
        if industry == "Semiconductors":
            return "semiconductor_value"
        if industry == "Semiconductor Equipment & Materials":
            return "semiconductor_equipment"
        if industry in {"Software - Application", "Software - Infrastructure"}:
            return "software_platform"
        if industry == "Information Technology Services":
            return "information_services"
        if industry == "Solar":
            return "energy_technology"
        return "technology_hardware"
    if sector == "Industrials":
        if industry == "Aerospace & Defense":
            return "defense_prime"
        if industry in {
            "Airlines",
            "Integrated Freight & Logistics",
            "Railroads",
            "Travel Services",
            "Trucking",
        }:
            return "transportation"
        if industry in {
            "Consulting Services",
            "Specialty Business Services",
            "Staffing & Employment Services",
        }:
            return "information_services"
        return "industrial_growth"
    raise RuntimeError(
        f"No explicit valuation profile for {ticker}: sector={sector!r}, industry={industry!r}"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--jansen-root", type=Path, default=DEFAULT_JANSEN_ROOT)
    parser.add_argument("--spy-holdings-xlsx", type=Path)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def cik_from_url(value: str | None) -> str | None:
    match = re.search(r"CIK=(\d+)", value or "", re.IGNORECASE)
    return match.group(1).zfill(10) if match else None


def active_ticker_metadata(root: Path) -> tuple[dict[str, dict], dict[str, list[dict]]]:
    columns = [
        "ticker",
        "name",
        "exchange",
        "isdelisted",
        "category",
        "cusips",
        "siccode",
        "sicsector",
        "sicindustry",
        "famaindustry",
        "sector",
        "industry",
        "currency",
        "location",
        "lastpricedate",
        "secfilings",
        "companysite",
    ]
    rows = ds.dataset(
        root / "tickers", format="parquet", partitioning="hive"
    ).to_table(columns=columns).to_pylist()
    by_ticker: dict[str, dict] = {}
    by_cusip: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        if row.get("isdelisted") != "N":
            continue
        ticker = str(row.get("ticker") or "").upper()
        if not ticker:
            continue
        current = by_ticker.get(ticker)
        if current is None or str(row.get("lastpricedate") or "") > str(
            current.get("lastpricedate") or ""
        ):
            by_ticker[ticker] = row
        for cusip in str(row.get("cusips") or "").split():
            by_cusip[cusip].append(row)
    return by_ticker, by_cusip


def current_jansen_members(root: Path) -> tuple[dt.date, list[dict]]:
    rows = ds.dataset(
        root / "sp500", format="parquet", partitioning="hive"
    ).to_table().to_pylist()
    snapshot_dates = [row["date"] for row in rows if row.get("action") == "current"]
    if not snapshot_dates:
        raise RuntimeError("Sharadar S&P 500 dataset has no current snapshot rows")
    as_of = max(snapshot_dates)
    members = [
        row for row in rows if row.get("action") == "current" and row["date"] == as_of
    ]
    return as_of, sorted(members, key=lambda row: row["ticker"])


def official_spy_holdings(path: Path | None) -> tuple[str | None, list[dict]]:
    if path is None:
        return None, []
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    as_of = None
    heading = worksheet.cell(3, 2).value
    if heading:
        match = re.search(r"(\d{1,2}-[A-Za-z]{3}-\d{4})", str(heading))
        if match:
            as_of = dt.datetime.strptime(match.group(1), "%d-%b-%Y").date().isoformat()
    rows = []
    for values in worksheet.iter_rows(min_row=6, values_only=True):
        name, ticker, cusip, _sedol, weight, _sector, shares, currency, *_ = values
        ticker = str(ticker or "").strip().upper()
        name = str(name or "").strip()
        if not ticker or ticker == "-" or name.startswith("CONTRA "):
            continue
        rows.append(
            {
                "ticker": ticker,
                "name": name,
                "cusip": str(cusip or "").strip(),
                "weightPct": float(weight) if weight is not None else None,
                "sharesHeld": float(shares) if shares is not None else None,
                "currency": str(currency or "").strip() or None,
            }
        )
    return as_of, rows


def available_source_tickers(root: Path, dataset_name: str, candidates: set[str]) -> set[str]:
    dataset = ds.dataset(root / dataset_name, format="parquet", partitioning="hive")
    filter_expression = ds.field("ticker").isin(sorted(candidates))
    if dataset_name == "fundamentals":
        filter_expression = filter_expression & ds.field("dimension").isin(["ARQ", "ART"])
    table = dataset.to_table(columns=["ticker"], filter=filter_expression)
    return {str(value).upper() for value in table.column("ticker").to_pylist() if value}


def main() -> None:
    args = parse_args()
    jansen_as_of, jansen_members = current_jansen_members(args.jansen_root)
    metadata_by_ticker, metadata_by_cusip = active_ticker_metadata(args.jansen_root)
    spy_as_of, spy_rows = official_spy_holdings(args.spy_holdings_xlsx)

    jansen_tickers = {str(row["ticker"]).upper() for row in jansen_members}
    spy_by_source: dict[str, dict] = {}
    unresolved_spy = []
    for row in spy_rows:
        source_ticker = row["ticker"] if row["ticker"] in jansen_tickers else None
        if source_ticker is None:
            matches_by_ticker = {
                str(candidate.get("ticker") or "").upper(): candidate
                for candidate in metadata_by_cusip.get(row["cusip"], [])
                if str(candidate.get("ticker") or "").upper() in jansen_tickers
            }
            matches = list(matches_by_ticker.values())
            if len(matches) == 1:
                source_ticker = str(matches[0]["ticker"]).upper()
        if source_ticker is None:
            unresolved_spy.append(row)
            continue
        spy_by_source[source_ticker] = row

    if spy_rows and unresolved_spy:
        raise RuntimeError(f"Unresolved official SPY holdings: {unresolved_spy}")
    if spy_rows and set(spy_by_source) != jansen_tickers:
        raise RuntimeError(
            "Official SPY and Sharadar current constituent sets differ after CUSIP mapping: "
            f"SPY-only={sorted(set(spy_by_source) - jansen_tickers)}, "
            f"Sharadar-only={sorted(jansen_tickers - set(spy_by_source))}"
        )

    candidate_tickers = {
        ticker
        for source_ticker in jansen_tickers
        for ticker in (
            source_ticker,
            (spy_by_source.get(source_ticker) or {}).get("ticker"),
        )
        if ticker
    }
    financial_tickers = available_source_tickers(
        args.jansen_root, "fundamentals", candidate_tickers
    )
    price_tickers = available_source_tickers(args.jansen_root, "prices", candidate_tickers)

    securities = []
    for member in jansen_members:
        source_ticker = str(member["ticker"]).upper()
        metadata = metadata_by_ticker.get(source_ticker)
        if metadata is None:
            raise RuntimeError(f"Missing active ticker metadata for {source_ticker}")
        cik = cik_from_url(metadata.get("secfilings"))
        if cik is None:
            raise RuntimeError(f"Missing CIK for current S&P 500 ticker {source_ticker}")
        official = spy_by_source.get(source_ticker, {})
        public_ticker = official.get("ticker") or source_ticker
        financial_ticker = (
            public_ticker if public_ticker in financial_tickers else source_ticker
        )
        price_ticker = public_ticker if public_ticker in price_tickers else source_ticker
        securities.append(
            {
                "sourceTicker": financial_ticker,
                "priceTicker": price_ticker,
                "membershipTicker": source_ticker,
                "publicTicker": public_ticker,
                "name": official.get("name") or member.get("name") or metadata.get("name"),
                "cusip": official.get("cusip")
                or str(metadata.get("cusips") or "").split()[0],
                "cik": cik,
                "spyWeightPct": official.get("weightPct"),
                "exchange": metadata.get("exchange"),
                "category": metadata.get("category"),
                "sector": metadata.get("sector"),
                "industry": metadata.get("industry"),
                "famaIndustry": metadata.get("famaindustry"),
                "sicCode": metadata.get("siccode"),
                "sicSector": metadata.get("sicsector"),
                "sicIndustry": metadata.get("sicindustry"),
                "currency": metadata.get("currency"),
                "location": metadata.get("location"),
                "lastPriceDate": metadata["lastpricedate"].isoformat()
                if metadata.get("lastpricedate")
                else None,
                "secFilings": metadata.get("secfilings"),
                "companySite": metadata.get("companysite"),
            }
        )

    by_cik: dict[str, list[dict]] = defaultdict(list)
    for security in securities:
        by_cik[security["cik"]].append(security)

    companies = []
    for cik, share_classes in sorted(by_cik.items()):
        ordered = sorted(
            share_classes,
            key=lambda row: (
                -(row["spyWeightPct"] if row["spyWeightPct"] is not None else -1),
                row["publicTicker"],
            ),
        )
        primary = ordered[0]
        aliases = sorted(
            {
                ticker
                for row in share_classes
                for ticker in (
                    row["publicTicker"],
                    row["sourceTicker"],
                    row["priceTicker"],
                    row["membershipTicker"],
                )
                if ticker != primary["publicTicker"]
            }
        )
        companies.append(
            {
                "ticker": primary["publicTicker"],
                "sourceTicker": primary["sourceTicker"],
                "aliases": aliases,
                "shareClasses": [
                    {
                        "ticker": row["publicTicker"],
                        "sourceTicker": row["sourceTicker"],
                        "priceTicker": row["priceTicker"],
                        "membershipTicker": row["membershipTicker"],
                        "cusip": row["cusip"],
                        "spyWeightPct": row["spyWeightPct"],
                    }
                    for row in ordered
                ],
                **{
                    key: primary[key]
                    for key in (
                        "name",
                        "cusip",
                        "cik",
                        "spyWeightPct",
                        "exchange",
                        "category",
                        "sector",
                        "industry",
                        "famaIndustry",
                        "sicCode",
                        "sicSector",
                        "sicIndustry",
                        "currency",
                        "location",
                        "lastPriceDate",
                        "secFilings",
                        "companySite",
                    )
                },
                "priceTicker": primary["priceTicker"],
                "membershipTicker": primary["membershipTicker"],
                "valuationProfile": valuation_profile(
                    primary["publicTicker"], primary.get("sector"), primary.get("industry")
                ),
            }
        )

    companies.sort(key=lambda row: row["ticker"])
    if len(securities) != 503 or len(companies) != 500:
        raise RuntimeError(
            f"Expected 503 securities / 500 issuers, got {len(securities)} / {len(companies)}"
        )

    payload = {
        "schemaVersion": 1,
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "asOf": spy_as_of or jansen_as_of.isoformat(),
        "securityCount": len(securities),
        "companyCount": len(companies),
        "dedupeKey": "SEC CIK",
        "primaryShareClassPolicy": "largest official SPY weight; alternate classes retained as aliases",
        "sources": {
            "officialSpyHoldings": OFFICIAL_HOLDINGS_URL if spy_rows else None,
            "officialSpyHoldingsAsOf": spy_as_of,
            "jansenSharadarSp500AsOf": jansen_as_of.isoformat(),
            "jansenSharadarDataset": "paid Sharadar parquet export",
        },
        "companies": companies,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(args.output),
                "asOf": payload["asOf"],
                "securityCount": len(securities),
                "companyCount": len(companies),
                "aliasCount": sum(len(row["aliases"]) for row in companies),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
